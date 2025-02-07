import openpyxl
from babel.dates import format_date

from django.conf import settings
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponse, HttpResponseRedirect
from django.db.models.functions import TruncDate
from django.db.models import Sum, Count, Subquery, OuterRef, Func, F

from oscar.core.loading import get_class, get_model

ReportGenerator = get_class("dashboard.reports.reports", "ReportGenerator")
ReportCSVFormatter = get_class("dashboard.reports.reports", "ReportCSVFormatter")
ReportHTMLFormatter = get_class("dashboard.reports.reports", "ReportHTMLFormatter")

Order = get_model("order", "Order")
Line = get_model("order", "Line")

file_path = settings.STATIC_PRIVATE_ROOT + "/xlsx/ОТЧЕТ!Прованс.xlsx"


class CRMReportCSVFormatter(ReportCSVFormatter):
    filename_template = "crm-%s-to-%s.csv"

    def generate_response(self, orders, **kwargs):
        if 0 == len(orders):
            msg = "Заказы за данный период не найдены!"
            messages.warning(
                kwargs["request"],
                msg,
            )
            return HttpResponseRedirect(reverse("dashboard:reports-index"))
        elif len(orders) > 31:
            msg = "Представленный диапазон для отчета больше 31 дня!"
            messages.warning(
                kwargs["request"],
                msg,
            )
            return HttpResponseRedirect(reverse("dashboard:reports-index"))
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # Найдем строку, после которой нужно вставить новые данные
            start_row = 16
            # Вставка данных из списка data в ячейки A16, B16, C16 и далее

            ws.cell(
                row=8,
                column=19,
                value=format_date(
                    orders.last().get("day"), format="LLLL", locale="ru"
                ).capitalize(),
            )
            ws.cell(row=10, column=19, value=orders.last().get("day").month)
            ws.cell(row=13, column=19, value=orders.last().get("day").year)

            for i, row_data in enumerate(orders):
                # ws.cell(row=start_row + i, column=1, value=row_data.get('day').replace(tzinfo=None))
                ws.cell(row=start_row + i, column=1, value=row_data.get("day"))
                ws.cell(row=start_row + i, column=22, value=row_data.get("total_sum"))
                ws.cell(row=start_row + i, column=28, value=row_data.get("order_count"))
                ws.cell(row=start_row + i, column=30, value=row_data.get("line_count"))

                ws.cell(row=start_row + i, column=6, value="-")
                ws.cell(row=start_row + i, column=11, value="-")
                ws.cell(row=start_row + i, column=17, value="-")

                ws.cell(row=start_row + i, column=33, value="не применимо")

            # Сохранение измененного файла
            file_name = "site-report.xlsx"
            wb.save(file_name)

            # Отправка файла в ответе
            with open(file_name, "rb") as file:
                response = HttpResponse(
                    file.read(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f"attachment; filename={file_name}"
                return response

    def filename(self, **kwargs):
        return self.filename_template % (kwargs["start_date"], kwargs["end_date"])


class CRMReportHTMLFormatter(ReportHTMLFormatter):
    filename_template = "oscar/dashboard/reports/partials/crm_report.html"


class CRMReportGenerator(ReportGenerator):
    code = "crm_report"
    description = "Отчет ТРЦ Планета"
    date_range_field_name = "day"
    queryset = (
        Order._default_manager.prefetch_related("lines")
        .filter(status=settings.OSCAR_SUCCESS_ORDER_STATUS)
        .annotate(
            day=TruncDate("date_placed"),
            # day=Func(
            #     F("date_placed"),
            #     function="DATE",  # SQLite поддерживает DATE функцию
            #     template="%(function)s(%(expressions)s)",
            # ),
            line_quantity=Subquery(
                Line.objects.filter(order=OuterRef("pk"))
                .values("order")
                .annotate(total_quantity=Sum("quantity"))
                .values("total_quantity")[:1]
            ),
        )
        .values("day")
        .annotate(
            order_count=Count("id", distinct=True),
            total_sum=Sum("total"),
            line_count=Sum("line_quantity"),
        )
        .order_by("day")
    )

    formatters = {
        "CSV_formatter": CRMReportCSVFormatter,
        "HTML_formatter": CRMReportHTMLFormatter,
    }

    def generate(self, *args, **kwargs):
        additional_data = {
            "start_date": self.start_date,
            "end_date": self.end_date,
            "request": args[0],
        }
        return self.formatter.generate_response(self.queryset, **additional_data)

    def is_available_to(self, user):
        return user.is_staff
